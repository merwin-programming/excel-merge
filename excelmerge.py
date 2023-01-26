# ceated by Merwin Dsouza (Jan 26, 2023)

from tkinter import Tk     # from tkinter import Tk for Python 3.x
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
from openpyxl import Workbook
from sys import exit
from os import system


def main():
    filepath, header_row, beg_col, end_col = init()
    data = getData(filepath, header_row, beg_col, end_col)
    data = format(data)
    filepath = filepath.removesuffix(".xlsx") + "_merged.xlsx"
    saveData(filepath, data)
    print("Success!")


def init():
    system("cls")
    Tk().withdraw()
    filepath = askopenfilename()
    if filepath == "":
        exit("Cancelled!")
    if not filepath.endswith(".xlsx"):
        exit("Invalid file!")

    try:
        header_row = int(input("Header row number (1, 2, 3, ...): "))
        if header_row < 1:
            raise ValueError()
    except ValueError:
        exit("Invalid row!")

    try:
        beg_col = input("Start column letter (A, B, C, ...): ").upper()
        if len(beg_col) != 1 or not beg_col.isalpha():
            raise ValueError()
    except ValueError:
        exit("Invalid column!")

    try:
        end_col = input("End column letter (A, B, C ...): ")
        if len(end_col) != 1 or not end_col.isalpha() or end_col < beg_col:
            raise ValueError()
    except ValueError:
        exit("Invalid column!")
        
    return filepath, header_row, beg_col, end_col


def getData(filepath, header_row, beg_col=None, end_col=None):
    beg_row = header_row + 1
    beg_col = ord(beg_col) - 64 # column letter to number (A = 1, B = 2, ...)
    end_col = ord(end_col) - 64

    data = []
    workbook = load_workbook(filepath)
    for sheet in workbook:
        for row in sheet.iter_rows(beg_row - 1 if len(data) == 0 else beg_row, sheet.max_row, beg_col, end_col, True):
            if row[0] == None:
                break
            data.append(row)
    workbook.close()
    
    return data


def saveData(filepath, data):
    workbook = Workbook()
    sheet = workbook.active
    for row in data:
        sheet.append(row)
    workbook.save(filepath)


def format(data):
    for i, row in enumerate(data):
        data[i] = list(row)
        data[i].insert(0, "Sr. No." if i == 0 else i)
        for j, value in enumerate(data[i]):
            if isinstance(value, str):
                data[i][j] = value.strip()
    
    return data
    

if __name__ == "__main__":
    main()